from openpyxl import *
wb = load_workbook('C:\\Users\\ghp\\Downloads\\Telegram Desktop\\Countries.xlsx')
names = wb.sheetnames
# print the available countries in the excel workbook
print(names)
country = input('please entre country name :')
# if the first letter of the country entered by the user is lowercase convert to uppercase to prevent any errors
if country.islower():
    country = country.title()
# country entered by the user not available he will enter another one
if country not in names:
    print("sorry this country is unavailable, the available countries are: ", names)
    country = input('please entre country name :')
ws = wb[str(country)]
print(ws.title)
print('what population do you want :\n entre 1 for population of the country \nentre 2 for population of each state '
      'in the country '
      '\n entre 3 for highest and lowest state population')
operation = int(input("Please entre number of operation you want"))
values = []
d = {}
x = 1
# if the user picked population of country, it append all cells in column 'B' and print its sum
if operation == 1:
    print("you have picked population of country")
    for row in ws['B']:
        values.append(row.value)
    print("total population of ", country, "is", sum(values))
# else if the user picked population of each state it prints each cell country under it its population
elif operation == 2:
    rows = ws.rows
    for row in rows:
        for cell in row:
            print(cell.value)
elif operation == 3:
    for cell in ws['A']:
        d[cell.value] = ' '
    for key in d:
        for cell in ws['B']:
            d.update(cell.value)
    min_population = min(d, key=d.get)
    print(min_population)
    print(d)
else:
    print("invalid input ")

